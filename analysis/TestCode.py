# -*- coding: utf-8 -*-
"""
Created on Mon Jul  6 12:21:13 2020

TEST CODE USED FOR PROCESSING ALL OF THE DATA FROM THE LARGE SCAN
THIS COLLECTS THE T60 and Absorption OF THE FULL SCAN

@author: cvong
"""

import numpy as np
#import byuarglib as byu
#import sys as directory
#keep using the natural working directory
#directory.path.insert(0,'C:/Users/cvongsaw/Box/UW Research/Code/uw-measurements')
#add a second working directory
#directory.path.insert(1,'C:/Users/cvongsaw/Box/UW Research/Code/underwater-measurements/analysis/')
from readLogFile import readLogFile
import matplotlib.pyplot as plt
import matplotlib.pylab as pylab
params = {'legend.fontsize': 24,
          'figure.figsize': (15, 10),
         'axes.labelsize': 28,
         'axes.titlesize':29,
         'axes.titleweight':'bold',
         'xtick.labelsize':24,
         'ytick.labelsize':24,
         'lines.linewidth':3}
pylab.rcParams.update(params)
import scipy.signal as sci 
import ESAUpose as pose
import ESAUdata as data
import ESAUResponse as response
import TankCharacterization as tank
import TimeGate_UnderwaterTank as tg
import xlsxwriter as xls
import openpyxl as xl
#import pdb
#pdb.set_trace()

###############################################################################
###############################################################################
#when switching file name via copy/paste, must change forward and back slashes
###############################################################################
###############################################################################
date4 = '2021-02-24' #scan 3,7,8(noise),9(noise)
date5 = '2021-02-26' #TL measurement [scan 3(5 points) and 4(10 points) good]
date6 = '2021-03-23' #short cal length long signal length
date11 = '2021-08-09'
date12 = 'Need to redo date12 good tests with panels in'
date13 = '2021-09-13'
date14 = '2021-09-14'
date15 = '2021-09-15'
date16 = '2021-09-16'
date17 = '2021-09-17B'
date18 = '2021-09-22'
date19 = '2021-09-27'
date2 = date19 #second file name
date = date2 #first file name

###NO PANELS####
#scan 2, 5k-10k 2021-09-14 downsample false [0.511,0.5155] octs 30
#scan7, 10k-50k 2021-09-15 downsample true factor2 [0.511,0.5149] octs 30
#scan 9, 50k-100k (24:729) 2021-09-15  scan 11 (0:24)  [0.5103,0.514] 2021-09-16 downsample factor2? octs 30
#scan 1, 100k-500k (1V) 2021-09-17B\2021-09-17 octs 25?
#scan 2, 100k-500k (3V) 2021-09-17B\2021-09-17 octs 25 (0:10)[0.6004,0.606]
###With PANELS####
#scan 3, 5k-10k 2021-09-27 octs 30 [0.6008,0.604]
#scan 4, 10k-50k 2021-09-27 octs 30 [0.6005,0.6025]
#scan 6, 50k-100k 2021-09-27 octs 30 [0.6005,0.6031]
#scan 1, 100k-500k (1V) 2021-09-22 octs 25 ?
#scan 4, 100k-500k (3V) 2021-09-22 octs 25 [0.6005,0.6035]
scan = '6'
octs = 30 #int(len(OctFreq)) #number of octave bands to look at
tbound = [0.6005,0.6031]
group = np.arange(50,100,1) #set of desired measurement IDs
downsample = False #may need to adjust factor
factor = 4 #factor by which to downsample
bandwidth = '50k-100k'
propagation = True #account for propagation absorption losses
water_air = False #account for minute impedance boundary of water-air surface
wall1 = "AbsorbAcrylic" #input data into appropriate spreadsheet name
wall2 = "AbsorbPanels" #input data into appropriate spreadsheet name
wall3 = "AbsorbAcrylicProp" #input data into appropriate spreadsheet name
wall4 = "AbsorbPanelsProp" #input data into appropriate spreadsheet name
walltype = wall4

for iii in group:
    print(f'Scan {iii} Calculation in progress...')
    #desire is the list of all the scans you care to actually look at in this analysis
    desire = [iii]#,19,56,80]
    filename = f'ID0{iii}_000log.txt'
    calfile = 'cal_000log.txt'
    channels = [0,1] #recording channels of interest
    cal_channel = [1] #recorded calibration of interest (currently should only be len(cal_channel)=1)
    #legend = ['0']
    #legend = ['Near Wall','Middle','Anechoic']
    
    year = date[0:4]
    xls_file = f'W:/Vongsawad/data/{date}/{walltype}{bandwidth}.xlsx'
    path = f'W:/Vongsawad/data/{date}/{date2[0:10]}_scan{scan}/'
    
    freqMin,freqMax,temp,fs,startzero,sigL,trail,totdur,depth,xSource,ySource,zSource,xRec,yRec,zRec = readLogFile(filename,path)
    _,_,_,_,_,_,_,_,_,xA_cal,yA_cal,zA_cal,xR_cal,yR_cal,zR_cal = readLogFile(calfile,path)
    temp = 20.88
    depth = 0.5
    
    N = fs*sigL #number of samples
    signal = f'{sigL}s Chirp {freqMin/1000}kHz-{freqMax/1000}kHz'
    test = f'{date} scan{scan} {signal}'
    
    #generated CALIBRATION signal
    fscal = fs #sampling frequency
    startzerocal = 0.02 #leading zeros of the signal
    sigLcal = sigL #signal length
    trailcal = 0.5 #trailing zeros
    treccal = sigLcal #time record length in sec
    Ncal = fscal*treccal
    Acal = (xA_cal,yA_cal,zA_cal)#(0.6,2.14,depth/2)
    Rcal = (xR_cal,yR_cal,zR_cal)#(0.6,2.06,depth/2)
    
    ###############################################################################
    ###############################################################################
    ##############################################################################
    ##############################################################################
    #load generated signal and calibration measurement
    #gen, calgen, cal, ch0, ch1, ch2, ch3 = data.ESAUdata(path+scan, desire, channels, N, Ncal)
    
    gen, _, cal, ch0, ch1, _, _ = data.ESAUdata(path, desire, channels, N, Ncal)
    calgen=gen
    
    #Need to downsample the signal for the T60 code to remove extra noise to IR**2 to at most 2.5x fmax
    if downsample == True:
        #resample the data for higher sampling rates???
        factor = factor #factor by which to reduce signal
        gen = sci.decimate(gen,factor) 
        calgen = sci.decimate(calgen,factor)
        ch0 = sci.decimate(ch0[:,0],factor)
        ch1 = sci.decimate(ch1[:,0],factor)
        fs = fs/factor #"""
        fscal = fs #sampling frequency
        caldec = np.empty([len(ch0),4])
        for i in range(4):
            caldec[:,i] = sci.decimate(cal[:,i],factor)
        cal = caldec
       
    #Load in positions, calculate range distance, plot scan positions desired
    A,R,dd = pose.ESAUpose(path, desire, plot=False, Acal = Acal, Rcal = Rcal)
    
    
    #time delays now allowing for a single measurement w/ single source/receiver pose
    c = tg.uwsoundspeed(D=depth,T=temp,S=0.03, model='Garrett')
    
    
    #need to update MeasGreen to handle various channels.and update the inputs
    #to use ch number instead of allsignals. 
    print('')
    print('')
    print('System Response...')
    cal1 = np.ndarray.flatten(cal[:,cal_channel]) #obtain only the cal of interest for calculating the system response
    #cal1 = cal[:,cal_channel[0]] ###!!change!!!####
    #cal1 = ch0[:,0]
    #ch1 = ch1[:,desire]
    #ch1 = ch1[:,0] ##!!change!!!####
    
    tgate,_,tdir,dis = tg.gateValue(A[0],R[0],c) 
    tgatec,_,tdirc,disc = tg.gateValue(Acal,Rcal,c) 
    
    #how much to gate the signal in samples
    tb4gate = 0.1 #ms before first reflection tgate
    Nb4gate = tb4gate/1000 *fs #convert to samples before gating
    Ngate = tgate*fs-Nb4gate
    
    tt = np.linspace(0,len(cal1)/fs,len(cal1))      #time array for recorded signal
    tt = tt*1000 
    """plt.figure()
    plt.plot(tt,ch1)
    plt.title(f'Recorded Signal \n {signal} fs={fs/1000}kHz \n {date} scan{scan}')
    plt.xlabel('time (ms)')
    plt.ylabel('Amplitude')"""
    
    hsys,tsys,Hsys,fsys = response.SysResponse(cal1,calgen,fscal,tgate=tgate,wiener=True,domain='f')
    if downsample == False:
        Htank,ftank = response.TankResponse(ch1[:,0],gen,fs,hsys,wiener=True,domain='f')
    if downsample == True:
        Htank,ftank = response.TankResponse(ch1,gen,fs,hsys,wiener=True,domain='f')
    htank = np.real(np.fft.ifft(Htank))
    ttank = np.linspace(0,len(htank)/fs,len(htank))
    
    
    """Hss = 2*Hsys[0:(int(len(Hsys)/2))]   #convert to single-sided FRF
    fss = fsys[0:int(len(fsys)/2)]/1000    #convert to single-sided from Hz to kHz
    Hss_dB = 10*np.log10(np.abs(Hss))   #convert to Levels (dB)
    
    plt.figure()
    plt.plot(fss,Hss_dB)
    plt.title(f'FRF of {signal} for System')
    plt.xlabel('Frequency (kHz)')
    plt.ylabel('Amplitude')
    plt.grid()
    
    plt.figure()
    plt.plot(tsys*1000,hsys)
    plt.axvline(tgatec*1000,linestyle='dashed',color='g')
    #plt.axvline(Ngate*1000,linestyle='dashdot',color='r') #no idea how to gate this relative to fs
    #need to compare this with the fs = 1M and fs = 10M measurements. 
    plt.axvline(tdirc*1000,linestyle='dashed',color='r')
    plt.xlabel('Time (ms)')
    plt.ylabel('Amplitude')
    plt.legend(['Chirp IR','Estimated First Reflection','Estimated Direct Signal'])
    plt.title(f'Time-Gated Calibration IR of {signal} fs ={fs}Hz')
    plt.xlim(0,0.8)
    plt.grid()"""
    
    
    
    
    
    """___Roll the Shape of the Time-Domain for htank___"""
    #The IR is shifted to the end of the array, such that the tail
    #spills over to the beginning of the ray. The array must be rolled
    #for alignment w/ the actual IR. However, the number of zeros must
    #be equal for both leading and trailing zeros. 
    roll = True
    if roll == True:
        rollt = int(0.5*len(htank)-1)
        shift = rollt/fs
        htank = np.roll(htank,rollt)
    else: 
        shift = 0
    """Htss = 2*Htank[0:(int(len(Htank)/2))]   #convert to single-sided FRF
    ftss = ftank[0:int(len(ftank)/2)]/1000    #convert to single-sided from Hz to kHz
    Htss_dB = 10*np.log10(np.abs(Htss))   #convert to Levels (dB)
    
    plt.figure()
    plt.plot(ftss,Htss_dB)
    plt.title(f'FRF of {signal} for Tank')
    plt.xlabel('Frequency (kHz)')
    plt.ylabel('Amplitude')
    plt.grid()
    #"""
    
    #####################################################
    """plt.figure()
    plt.plot(ttank,htank)
    plt.axvline(tgate+shift,linestyle='dashed',color='g')
    #plt.axvline(Ngate*1000,linestyle='dashdot',color='r') #no idea how to gate this relative to fs
    #need to compare this with the fs = 1M and fs = 10M measurements. 
    plt.axvline(tdir+shift,linestyle='dashed',color='r')
    plt.xlabel('Time (s)')
    plt.ylabel('Amplitude')
    plt.legend(['Chirp IR','Estimated First Reflection','Estimated Direct Signal'])
    plt.title(f'Calibrated IR of {signal} fs ={fs}Hz')
    #plt.xlim(0,3)
    plt.grid()"""
    
    
    #tbound = tank.T60meas_bounds(htank,fs)
    T60 = tank.T60meas(htank,fs,tbound[0],tbound[1],d=depth,c=c,rt='T10',plot=False)
    
    octData,OctFreq = tank.OctaveFilter(htank,freqMin,freqMax,fs,frac=octs)
    octTrans = np.transpose(octData)
    
    if propagation == True:
        #propagation absorption estimated for the water characteristics over desired Octave Bands
        #prop=0 #when desired to not look into propagation effects. 
        prop = tank.alpha_prop(OctFreq,T=temp,S=5,pH=7.2,depth=depth) 
    else: prop = np.zeros(len(OctFreq))
    
    """plt.figure()
    legend1 = []
    for i in range(len(OctFreq)):
        plt.plot(ttank,octData[i,:])
        legend1.append(np.round(OctFreq[i]))
    plt.xlabel('Time (s)')
    plt.ylabel('Amplitude')
    plt.title('IR Octave Band')
    plt.legend(legend1)"""
    
    """plt.figure()
    legend = []
    for i in range(len(OctFreq)):
        plt.plot(ttank,octTrans[:,i])
        legend.append(np.round(OctFreq[i]))
    plt.xlabel('Time (s)')
    plt.ylabel('Amplitude')
    plt.title('IR Octave Band decimated')
    plt.legend(legend1)"""
    
    
    T60_f = np.empty(int(len(OctFreq)))
    a_wall_f = np.empty(int(len(OctFreq)))
    for i in range(len(OctFreq)):
        #tbound = tank.T60meas_bounds(octTrans[:,i],fs)
        T60_f[i] = tank.T60meas(octTrans[:,i],fs,tbound[0],tbound[1],d=depth,c=c,rt='T10',plot=False)
        #plt.suptitle(f'f = {OctFreq[i]/1000} kHz')
        a_wall_f[i] = tank.alpha_wall(T60_f[i],d=depth,c=c,acc=water_air,alpha_p=prop[i])
        
    #overall absorption coefficient over entire bandwidth. Cannot include 
    #propagation absortion which is freq. dependent for a full scan bandwidth
    a_wall_gen = tank.alpha_wall(T60,d=depth,c=c,acc=water_air,alpha_p=0)
    T60est,sig,fschroeder = tank.T60est(depth,c=c,)
    
    
    
    #Open an excel file and append data to it to create a repository of data
    #this is particularly important for sharing absorption coefficient values
    #to apply to the models Kaylyn is working on. 
    #https://realpython.com/openpyxl-excel-spreadsheets-python/
    #https://openpyxl.readthedocs.io/en/stable/tutorial.html
    #https://www.geeksforgeeks.org/python-writing-excel-file-using-openpyxl-module/
    wb = xl.load_workbook(xls_file)     #Find created worksheet
    a_sheet = wb['Alpha']               #Find worksheet names
    T_sheet = wb['T60']
    #populate selected excel workbook with alpha (absorption coefficient data)
    for i in range(len(OctFreq)):
        c1 = a_sheet.cell(row=1,column=i+3) 
        c1.value = OctFreq[i]               #populate alpha frequencies
        c1 = T_sheet.cell(row=1,column=i+3) 
        c1.value = OctFreq[i]               #populate T60 frequencies
    for i in range(len(a_wall_f)):
        c01 = a_sheet.cell(row=desire[0]+3,column = 2)
        c01.value = a_wall_gen              #populate overall alpha
        c2 = a_sheet.cell(row=desire[0]+3,column=1)
        c2.value = f'alpha_{desire[0]}'     #order alpha values by ID#
        c3 = a_sheet.cell(row=desire[0]+3,column=i+3)
        c3.value = a_wall_f[i]              #populate alpha values by freq bin
    
    #populate selected excel workbook with T60 (Reverberation time data)
    for i in range(len(a_wall_f)):
        c01 = T_sheet.cell(row=desire[0]+3,column = 2)
        c01.value = T60                     #populate overall T60 value
        c2 = T_sheet.cell(row=desire[0]+3,column=1)
        c2.value = f'T60_{desire[0]}'       #order T60 values by ID#
        c3 = T_sheet.cell(row=desire[0]+3,column=i+3)
        c3.value = T60_f[i]                 #populate T60 values by freq bin
    wb.save(xls_file)   #save the updated workbook



"""
ht = np.abs(np.fft.ifft(Htank))
T60 = tank.T60meas(ht,fs,t0=0,t1=int(len(ht)/fs),d=depth,c=c,rt='T60',plot=True)
alpha_s = tank.alpha_wall(T60,depth,c,acc=False,anech=False,alpha_p=0)
"""
"""
Hss = 2*Hsys[0:(int(len(Hsys)/2))]   #convert to single-sided FRF
fss = fsys[0:int(len(fsys)/2)]/1000    #convert to single-sided from Hz to kHz
Hss_dB = 10*np.log10(Hss) 

Hsstank = 2*Htank[0:(int(len(Htank)/2))]   #convert to single-sided FRF
fsstank = ftank[0:int(len(ftank)/2)]/1000    #convert to single-sided from Hz to kHz
Hsstank_dB = 10*np.log10(Hsstank/1e-6) 
"""
"""
ir = response.IR(ch1,gen,fs,wiener=False,domain='f')
ir1 = 10*np.log10(ir**2)
tt = np.linspace(0,len(ir1)/fs,len(ir1))               #time array for ht
tt = tt*1000 
frf = np.fft.fft(ir)
frf = 10*np.log10(frf/1e-6)
frf = 2*frf[0:(int(len(frf)/2))]
f = np.fft.fftfreq(len(ir),d=1/fs)
f = f[0:int(len(f)/2)]/1000
plt.figure()
plt.plot(tt,ir)
plt.title('straight IR')
plt.xlabel('time (ms)')
plt.ylabel('')
plt.grid()
plt.figure()
plt.plot(f,frf)
plt.title(f'Straight FRF')
plt.xlabel('Frequency (kHz)')
plt.ylabel(r'Level (dB re 1 $\mu Pa$)')
plt.grid()
"""




"""
plt.figure()
plt.plot(fss,Hss_dB)
plt.title(f'Time-Gated Calibration FRF of {signal} Swept-Sine')
plt.xlabel('Frequency (kHz)')
plt.ylabel(r'Level (dB re 1 $\mu Pa$)')
#plt.xlim(0,300)
#plt.ylim(35,65)
plt.grid()

plt.figure()
plt.plot(fsstank,Hsstank_dB)
plt.title(f'Calibrated Tank Transfer Function of {signal} Swept-Sine')
plt.xlabel('Frequency (kHz)')
plt.ylabel(r'Level (dB re 1 $\mu Pa$)')
#plt.xlim(0,300)
#plt.ylim(35,65)
plt.grid()
"""







"""
#need to EDIT with the below link
#https://www.askpython.com/python/examples/rmse-root-mean-square-error

"""


##############################################################################
##############################################################################
#OASPL w/ and w/out panels
##############################################################################
##############################################################################

##############################################################################
"""OASPL"""
##############################################################################
"""
#Calculate the overall sound pressure level with respect to 1 microPascal as is
#standard for underwater acoustics, rounded to 4 decimal places. 
print('')
print('calculating OASPL re 1e-6 Pa from timewaveform...')
OASPL = np.empty(len(desire))
for i in range(len(desire)):
    OASPL[i] = np.round(10 * np.log10( np.mean( np.square( ch0[:,i] ) ) /1e-6**2),4)
print(f'OASPL{legend} = {OASPL}')
"""
##############################################################################
"""
gate_dir = np.argwhere(t<0.34e-3)
gate_dir = max(gate_dir[:,0])
gate_min = np.argwhere(t>0.34e-3)
gate_min = min(gate_min[:,0])
gate_max = np.argwhere(t<0.40e-3)
gate_max = max(gate_max[:,0])
IR_gate_dir = np.real(x0[:gate_dir, 0])
IR_gate_anech = np.real(x0[gate_min:gate_max, 2])
IR_gate_acryl = np.real(x0[gate_min:gate_max, 0])

#OASPL 
OASPL = np.empty((1,len(desire)))
for i in range(len(desire)):
    OASPL[:,i] = byu.OASPLcalc(ch0[:,i])
print(OASPL)
"""



##############################################################################
##############################################################################

"""
#Calculate the overall sound pressure level with respect to 1 microPascal as is
#standard for underwater acoustics, rounded to 4 decimal places. This is
#calculating from the IR instead of the timewaveform as seen above. 
print('')
print('calculating IR OASPL re 1e-6 Pa...')
OASPL = np.empty(len(desire))
for i in range(len(desire)):
    OASPL[i] = np.round(10 * np.log10( np.mean( np.square( x0[:,i] ) ) /1e-6**2),4)
print(f'OASPL{legend} = {OASPL}')
"""

##############################################################################
##############################################################################